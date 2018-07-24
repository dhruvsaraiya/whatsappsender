import openpyxl as xl
import string

def openWorkbook(filename, sheetname):
    wb = xl.load_workbook(filename)
    print wb.sheetnames
    if sheetname in wb.sheetnames:
        ws = wb[sheetname]
        return ws
    else:
        print "Sheet " + sheetname + "does not exist"
        print "Returning active sheet"
        return wb.active


#ws = openWorkbook(filename, sheetname)

# def getColumns(ws, columns):
#     col_n = list()
#     for col_name in columns:
#         flag = 0
#         for i, col in enumerate(ws.iter_cols()):
#             cell_val = str(ws.cell(row=1, column = i+1).value)
#             #print type(cell_val)
#             #print cell_val
#             if cell_val == col_name:
#                 print col_name, ' found - ', i+1
#                 col_n.append(i+1)
#                 flag = 1
#                 break
#         if flag==0:
#             print col_name, ' not found'
#     return col_n

# def getAllColumns():
#     first_row = list(ws.rows)[0]
#     col_list = list()
#     for cell in first_row:
#         col_list.append(cell.value)
#     return col_list

def getRowFromIndex(row):
    # row = list(ws.rows)[index]
    values = list()
    for cell in row:
        values.append(cell.value)
    return values

def getRows(ws):
    # col_n = getAllColumns(ws)
    rows_list = list(ws.rows)
    col_names = getRowFromIndex(rows_list[0])
    values_list = list()
    for i in range(1, len(rows_list)):
        row = getRowFromIndex(rows_list[i])
        row_dict = {}
        for index in range(len(col_names)):
            if row[index] is not None:
                try:
                    row_dict[col_names[index]] = str(row[index])
                except Exception as err:
                    print(err)

            else:
                row_dict[col_names[index]] = None
        values_list.append(row_dict)
    return col_names, values_list


# def getMessages(ws):
#     #col_n = getColumns(ws, column_names)
#     col_n = getAllColumns(ws)

#     # col_n.sort()
#     ##first name, second number, third message
#     n_rows = ws.max_row
#     print "total_messages ", n_rows-1

#     messages = list()

#     for i in xrange(n_rows-1):
#         user_details = list()
#         for j, n in enumerate(col_n):
#             val = ws.cell(row=i+2, column = n).value
#             if isinstance(val, float):
#                 val = int(val)
#                 val = str(val)
#             '''
#             if j==2:
#                 #val = val.encode('utf-8')
#                 #val = val.replace(u'', u'')
#                 val = "*Greetings from LogIQids*. Please note that we have posted Aadhya Mutha certificate via Speed Post. Please let us know if you do not receive the same by end of this week."
#                 #print val
#             else:
#                 val = str(val)
#                 #print val
#             '''
#             user_details.append(val)
#         messages.append(user_details)
#     return messages
