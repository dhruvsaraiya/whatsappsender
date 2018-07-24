import sys
import os
from PyQt4 import QtGui, QtCore
# import backend_driver
# from multiprocessing import Process, Value, freeze_support
import time
import get_messages_v2 as gm
import datetime

import openpyxl as xl
from openpyxl import Workbook
from string import Formatter
import pyqrcode

sys.stdout = open('gui_output.txt', 'w')
sys.stderr = open('gui_error.txt', 'w')

sheetname = "Sheet1"
w_percentage = 75

class NaughtyThread(QtCore.QThread):

    def __init__(self, win):
        QtCore.QThread.__init__(self)
        self.win = win

    def run(self):
        while(True):
            # print "Status : ", self.win.proc.is_alive()
            if self.win.go_process.state() == QtCore.QProcess.Running:
                print("Running")
                qrstring = ""
                qrimage = 'qrcode.png'
                if os.path.exists(self.win.qr_file):
                    with open(self.win.qr_file) as f:
                        qrstring = f.read()
                    qrcode = pyqrcode.create(qrstring)
                    if os.path.exists(qrimage):
                        os.remove(qrimage)
                    with open(qrimage, 'wb') as fstream:
                        qrcode.png(fstream, scale=5)
                    self.emit(QtCore.SIGNAL("showImage( PyQt_PyObject )"), qrimage)
                    time.sleep(10)
                    os.remove(self.win.qr_file)
                    os.remove(qrimage)
                else:
                    self.emit(QtCore.SIGNAL("showImage( PyQt_PyObject )"), None)

                try:
                    with open('status.txt', 'r') as f:
                        s = f.read()
                    status = "Completed : " + str(s) + " %"
                    self.emit(QtCore.SIGNAL("setStatus( PyQt_PyObject )"), status)

                except Exception as err:
                    print(err)
                err_flag = False
                time.sleep(3)
            else:
                err_flag = False
                if os.path.exists('my_err.txt'):
                    with open('my_err.txt') as f:
                        errstring = f.read()
                    if errstring == "restore":
                        import tempfile
                        err_flag = True
                        tempdir = tempfile.gettempdir()
                        session_file = os.path.join(tempdir, "whatsappSession.gob")
                        if os.path.exists(session_file):
                            os.remove(session_file)
                    #if errstring == "qrcode":
                        err_flag = True
                    self.emit(QtCore.SIGNAL("reRun( PyQt_PyObject )"), errstring)
                    self.emit(QtCore.SIGNAL("setError( PyQt_PyObject )"), "Scan QR Code")
                    os.remove('my_err.txt')

                try:
                    with open('status.txt', 'r') as f:
                        s = f.read()
                    status = "Completed : " + str(s) + " %"
                    self.emit(QtCore.SIGNAL("setStatus( PyQt_PyObject )"), status)

                except Exception as err:
                    print(err)
                print("End")
                if not err_flag:
                    break
        return


# class QrPopup(QtGui.QWidget):

#     def __init__(self, parent):
#         super(QrPopup, self).__init__(parent)
#         qrimage = 'qrcode.png'
#         self.qr_image = QtGui.QLabel(self)
#         self.qr_image.move(10, 10)
#         pixmap = QtGui.QPixmap(qrimage)
#         pixmap = pixmap.scaled(50, 50, QtCore.Qt.KeepAspectRatio)
#         self.qr_image.setPixmap(pixmap)


class Window(QtGui.QWidget):

    def __init__(self):
        super(Window, self).__init__()
        self.excel_filename = None
        self.image_filename = None
        self.qr_file = "qrcode.txt"
        self.options = ['Message First', 'Image First', 'As Caption']
        # self.clear = Value('i', 0)
        # self.status = Value('i', 0)
        # self.proc = None
        self.notifier = None
        self.excel_data = None
        self.excel_approve_flag = False
        self.initDimensions()
        self.initUI()
        self.time_stamp = None
        self.go_process = QtCore.QProcess(self)
        # self.go_process.started.connect(self.go_started)
        self.go_process.finished.connect(self.go_finished)
        self.qr_popup = None

    def initDimensions(self):
        # dim = QtGui.QDesktopWidget.screenGeometry()
        dim = QtGui.QDesktopWidget().screenGeometry()
        print (dim.width(), dim.height())
        # global window_width,window_height
        self.window_width = dim.width() * w_percentage / 100
        self.window_height = dim.height() * w_percentage / 100

        h_ten = self.window_height * 0.01321
        self.h_ten = h_ten
        h_twenty = self.window_height * 0.026
        h_thirty = self.window_height * 0.039
        h_fourty = self.window_height * 0.052
        h_hundred = self.window_height * 0.1321

        w_ten = self.window_width * 0.00732064
        w_twenty = self.window_width * 0.01464
        w_thirty = self.window_width * 0.02196
        w_fourty = self.window_width * 0.02928
        w_fifty = self.window_width * 0.0366
        w_hundred = self.window_width * 0.0732064


        self.uploaded_image_width = 1.5 * w_hundred
        self.uploaded_image_height = 1.5 * h_hundred
        self.logo_image_width = 180
        self.logo_image_height = 50
        self.logo_image_x = self.window_width / 2 - self.logo_image_width / 2
        self.logo_image_y = 0

        self.left_x = w_ten
        self.right_x = self.window_width / 2 + h_ten
        self.left_start_y = self.logo_image_height + self.logo_image_y
        self.right_start_y = self.logo_image_height + self.logo_image_y

        self.line_start = self.left_start_y

        self.excel_table_width = self.right_x - h_thirty
        self.excel_table_height = 3 * h_hundred

        self.text_area_width = (self.window_width - self.right_x) - w_twenty
        self.text_area_height = 1.5 * h_hundred


        self.error_label_x = self.right_x + self.logo_image_width / 2 + w_ten
        self.error_label_y = self.right_start_y / 2 - h_twenty
        self.status_label_x = self.left_x
        self.status_label_y = self.left_start_y / 2

        # clear_button_x = 200
        # clear_button_y = 40

        self.excel_upload_button_x = (0 + self.right_x) / 2 - (5 * w_ten)
        self.excel_upload_button_y = self.left_start_y + h_twenty

        self.excel_table_x = self.left_x
        self.excel_table_y = self.excel_upload_button_y + (5 * h_ten)
        self.excel_approve_button_x = self.excel_upload_button_x / 2
        self.excel_approve_button_y = self.excel_table_y + self.excel_table_height + h_twenty
        self.excel_cancel_button_x = self.excel_upload_button_x + (self.excel_upload_button_x / 2)
        self.excel_cancel_button_y = self.excel_table_y + self.excel_table_height + h_twenty

        self.qr_image_x = self.excel_upload_button_x - 15 * h_ten
        self.qr_image_y = self.excel_cancel_button_y + 3 * h_ten

        self.text_area_label_x = self.right_x
        self.text_area_label_y = self.right_start_y + h_thirty
        self.text_area_x = self.right_x
        self.text_area_y = self.text_area_label_y + (4 * h_ten)

        self.image_upload_button_x = (self.window_width - self.right_x) / 2 + self.right_x - (5 * w_ten)
        self.image_upload_button_y = self.text_area_y + self.text_area_height + h_twenty
        self.remove_image_button_x = self.image_upload_button_x
        self.remove_image_button_y = self.image_upload_button_y
        # self.uploaded_image_x = self.image_upload_button_x - w_twenty
        self.uploaded_image_x = (self.window_width - self.right_x) + (10 * w_ten)
        self.uploaded_image_y = self.image_upload_button_y + (5 * h_ten)

        self.option_buttons_start_x = (self.window_width - self.right_x) / 2 + self.right_x
        self.option_buttons_start_y = self.uploaded_image_y

        self.send_message_button_x = self.image_upload_button_x
        # send_message_button_y = uploaded_image_y + uploaded_image_height + 10
        self.send_message_button_y = self.image_upload_button_y + (5 * h_ten)
        # self.connected_button_x = self.send_message_button_x
        # self.connected_button_y = self.send_message_button_y

        self.error_link_x = self.error_label_x
        self.error_link_y = self.error_label_y + h_thirty

        self.window_title = "WhatsApp Sender"
        self.icon_image = "important/whatsapp.png"
        self.logiqids_logo = "important/logiqids.png"

    def closeEvent(self, event):
        try:
            if self.go_process.state() == QtCore.QProcess.Running:
                self.go_process.kill()
        except Exception as err:
            print(err)
        print ("closed")

    def initUI(self):
        #self.setGeometry(window_offset_x, window_offset_y, window_width, window_height)
        self.resize(self.window_width, self.window_height)
        self.center()

        # background color
        p = self.palette()
        p.setColor(self.backgroundRole(), QtGui.QColor("white"))
        self.setPalette(p)

        # logiqids logo
        self.logo_image = QtGui.QLabel(self)
        self.logo_image.move(self.logo_image_x, self.logo_image_y)
        pixmap = QtGui.QPixmap(self.logiqids_logo)
        pixmap = pixmap.scaled(self.logo_image_width, self.logo_image_height, QtCore.Qt.KeepAspectRatio)
        # pixmap.scaledToHeight(uploaded_image_height)
        self.logo_image.setPixmap(pixmap)

        # self.logo_image.resize(uploaded_image_width, uploaded_image_height)

        # error label
        self.error_label = QtGui.QLabel("", self)
        self.error_label.move(self.error_label_x, self.error_label_y)
        self.error_label.setStyleSheet('color: red')

        # status label
        self.status_label = QtGui.QLabel("", self)
        self.status_label.move(self.status_label_x, self.status_label_y)
        self.status_label.setStyleSheet('color: green')

        # # Clear Button
        # self.clear_button = QtGui.QPushButton('Clear', self)
        # self.clear_button.setToolTip('Clear WhatsApp Search Box and hit Me!!!!')
        # self.clear_button.move(clear_button_x, clear_button_y)
        # self.clear_button.resize(self.clear_button.sizeHint())
        # self.clear_button.clicked.connect(self.clear_button_clicked)
        # self.clear_button.setVisible(False)


        # Excel Upload Button
        self.excel_upload_button = QtGui.QPushButton('Upload Excel', self)
        self.excel_upload_button.setToolTip('Upload Excel Sheet')
        self.excel_upload_button.move(self.excel_upload_button_x, self.excel_upload_button_y)
        self.excel_upload_button.resize(self.excel_upload_button.sizeHint())
        self.excel_upload_button.clicked.connect(self.excel_upload_button_clicked)


        # Excel Table
        self.excel_table = QtGui.QTableWidget(self)
        self.excel_table.move(self.excel_table_x, self.excel_table_y)
        self.excel_table.setVisible(False)
        self.excel_table.setEditTriggers(QtGui.QAbstractItemView.NoEditTriggers)
        self.excel_table.resize(self.excel_table_width, self.excel_table_height)

        # Excel Approve Button
        self.excel_approve_button = QtGui.QPushButton('Approve Excel', self)
        self.excel_approve_button.setToolTip('Approve uploaded excel sheet')
        self.excel_approve_button.move(self.excel_approve_button_x, self.excel_approve_button_y)
        self.excel_approve_button.resize(self.excel_approve_button.sizeHint())
        self.excel_approve_button.clicked.connect(self.excel_approve_button_clicked)
        self.excel_approve_button.setVisible(False)

        # Excel Cancel Button
        self.excel_cancel_button = QtGui.QPushButton('Cancel', self)
        self.excel_cancel_button.setToolTip('Retry again!!')
        self.excel_cancel_button.move(self.excel_cancel_button_x, self.excel_cancel_button_y)
        self.excel_cancel_button.resize(self.excel_cancel_button.sizeHint())
        self.excel_cancel_button.clicked.connect(self.excel_cancel_button_clicked)
        self.excel_cancel_button.setVisible(False)

        #QR Code Image
        self.qr_image = QtGui.QLabel(self)
        self.qr_image.move(self.qr_image_x, self.qr_image_y)
        self.qr_image.resize(200, 200)
        self.qr_image.setVisible(False)

        # Text Area Label
        self.text_area_label = QtGui.QLabel("Message : ", self)
        self.text_area_label.move(self.text_area_label_x, self.text_area_label_y)

        # Text Area
        self.text_area = QtGui.QPlainTextEdit(self)
        self.text_area.resize(self.text_area_width, self.text_area_height)
        self.text_area.move(self.text_area_x, self.text_area_y)

        # Image Upload Button
        self.image_upload_button = QtGui.QPushButton('Upload Image', self)
        self.image_upload_button.setToolTip('Upload Picture to be Sent to Parents')
        self.image_upload_button.move(self.image_upload_button_x, self.image_upload_button_y)
        self.image_upload_button.resize(self.image_upload_button.sizeHint())
        self.image_upload_button.clicked.connect(self.image_upload_button_clicked)

        # Uploaded Image
        self.uploaded_image = QtGui.QLabel(self)
        self.uploaded_image.move(self.uploaded_image_x, self.uploaded_image_y)
        self.uploaded_image.resize(self.uploaded_image_width, self.uploaded_image_height)
        self.uploaded_image.setVisible(False)

        # Remove Image Button
        self.remove_image_button = QtGui.QPushButton('Remove Image', self)
        self.remove_image_button.setToolTip('Remove Uploaded Image')
        self.remove_image_button.move(self.remove_image_button_x, self.remove_image_button_y)
        self.remove_image_button.resize(self.remove_image_button.sizeHint())
        self.remove_image_button.clicked.connect(self.remove_image_button_clicked)
        self.remove_image_button.setVisible(False)

        # Options' Radio Buttons
        self.option_button_group = QtGui.QButtonGroup(self)
        self.option_buttons = []

        for i in range(len(self.options)):
            self.option_buttons.append(QtGui.QRadioButton(self.options[i], self))
            self.option_buttons[i].move(self.option_buttons_start_x, self.option_buttons_start_y + (self.h_ten * i * 3))
            self.option_buttons[i].setVisible(False)
            self.option_button_group.addButton(self.option_buttons[i], i)
            self.connect(self.option_buttons[i], QtCore.SIGNAL("clicked()"), self.option_button_clicked)
        self.selected_option = 0
        self.option_buttons[self.selected_option].setChecked(True)


        # Send Message Button
        self.send_message_button = QtGui.QPushButton('Send Message(s)', self)
        self.send_message_button.setToolTip('Send Messages Via WhatsApp')
        self.send_message_button.move(self.send_message_button_x, self.send_message_button_y)
        self.send_message_button.resize(self.send_message_button.sizeHint())
        self.send_message_button.clicked.connect(self.send_message_button_clicked)

        # # Connected Button
        # self.connected_button = QtGui.QPushButton('Connected', self)
        # self.connected_button.setToolTip('Click when Connected to WhatsApp')
        # self.connected_button.move(self.connected_button_x, self.connected_button_y)
        # self.connected_button.resize(self.connected_button.sizeHint())
        # self.connected_button.clicked.connect(self.connected_button_clicked)
        # self.connected_button.setVisible(False)

        # Notifier Thread
        self.notifier = NaughtyThread(self)
        QtCore.QObject.connect(self.notifier, QtCore.SIGNAL("showImage( PyQt_PyObject )"), self.show_image_from_thread)
        QtCore.QObject.connect(self.notifier, QtCore.SIGNAL("setStatus( PyQt_PyObject )"), self.set_status_from_thread)
        QtCore.QObject.connect(self.notifier, QtCore.SIGNAL("setError( PyQt_PyObject )"), self.set_error_from_thread)
        QtCore.QObject.connect(self.notifier, QtCore.SIGNAL("reRun( PyQt_PyObject )"), self.re_run_from_thread)

        # QtCore.QObject.connect(self.notifier, QtCore.SIGNAL("setConnected( PyQt_PyObject )"), self.set_connected_from_thread)
        self.notifier.finished.connect(self.n_finish)
        # self.first_attempt = True

        # Error File Link
        self.tb_label = QtGui.QLabel(self)
        self.tb_label.move(self.error_link_x, self.error_link_y - 10)
        self.tb_label.setText("Error File : ")
        self.tb_label.setVisible(False)
        self.tb = QtGui.QLabel(self)
        self.tb.setOpenExternalLinks(True)
        newfont = QtGui.QFont("Times", 12, QtGui.QFont.Normal)
        self.tb.move(self.error_link_x, self.error_link_y)
        self.tb.setFont(newfont)
        self.err_dir = os.path.expanduser("~\Desktop")
        self.error_file = os.path.join(self.err_dir, "error.xlsx")
        self.tb.setText('<a target="_blank" href="file:///{path}">{path}</a>'.format(path=self.error_file))
        self.tb.setVisible(False)

        self.display_side_panel(False)
        self.setWindowTitle(self.window_title)
        self.setWindowIcon(QtGui.QIcon(self.icon_image))
        self.show()


    # def clear_button_clicked(self):
    #     backend_driver.setClear(0)
    #     self.clear_button.setVisible(False)

    def option_button_clicked(self):
        self.selected_option = self.option_button_group.checkedId() + 1
        # print(self.option_button_group.checkedId())
        # print(self.option_button_group.checkedButton().text())


    def paintEvent(self, event):
        # line drawing
        paint = QtGui.QPainter()
        paint.begin(self)
        paint.drawLine(0, self.line_start, self.window_width, self.line_start)
        paint.drawLine(self.window_width / 2, self.line_start, self.window_width / 2, self.window_height)
        # paint.save()

    def set_status_from_thread(self, value):
        self.set_status(value, "green")

    def set_error_from_thread(self, value):
        self.set_error(value, "red")

    # def show_image_from_thread(self, qrimage):
    #     import matplotlib.pyplot as plt
    #     import matplotlib.image as mpimg
    #     import numpy as np
    #     print("inside........of qr image showing.")
    #     img = mpimg.imread(qrimage)
    #     imgplot = plt.imshow(img)
    #     plt.show()
    def show_image_from_thread(self, qrimage):
        if qrimage is None:
            self.qr_image.setVisible(False)
            return
        pixmap = QtGui.QPixmap(qrimage)
        pixmap = pixmap.scaled(200, 200, QtCore.Qt.KeepAspectRatio)
        self.qr_image.setPixmap(pixmap)
        self.qr_image.setVisible(True)
        print("qr must be displayed....")
        # pixmap = pixmap.scaled(self.uploaded_image_width, self.uploaded_image_height)

    def re_run_from_thread(self, error):
        try:
            if os.path.exists(self.qr_file):
                os.remove(self.qr_file)
            if os.path.exists("status.txt"):
                os.remove("status.txt")
        except Exception as err:
            print(err)

        go_exe = "whatsapp.exe"
        go_file = "whatsapp.xlsx"
        if self.image_filename:
            arguments = [go_file, self.image_filename, str(self.selected_option), self.time_stamp]
        else:
            arguments = [go_file, self.time_stamp]
        self.go_process.start(go_exe, arguments)
    # def set_connected_from_thread(self, value):
    #     backend_driver.setFlag_2(self.connected, value)
    #     self.connected_button.setEnabled(True)

    def remove_image_button_clicked(self):
        self.image_filename = None
        self.uploaded_image.setVisible(False)
        self.remove_image_button.setVisible(False)
        self.image_upload_button.setVisible(True)
        self.send_message_button_y = self.image_upload_button_y + (5 * self.h_ten)
        self.send_message_button.move(self.send_message_button_x, self.send_message_button_y)
        # self.connected_button_y = self.send_message_button_y
        # self.connected_button.move(self.connected_button_x, self.connected_button_y)
        for btn in self.option_buttons:
            btn.setEnabled(False)
            btn.setVisible(False)


    def image_upload_button_clicked(self):
        self.image_filename = QtGui.QFileDialog.getOpenFileName(self)
        self.image_filename = str(self.image_filename)
        base = os.path.basename(self.image_filename)
        file, ext = os.path.splitext(base)
        if (ext != '.png' and ext != '.jpg' and ext != '.jpeg'):
            print ("Invalid Image")
            # l.resize(self.upload_button.frameGeometry().width(), self.upload_button.frameGeometry().height())
            # l.move(self.upload_button.frameGeometry().width() + upload_button_x + 10 , upload_button_y)
            self.set_error("Not Valid Image", "red")
            self.image_filename = None
        else:
            pixmap = QtGui.QPixmap(self.image_filename)
            # pixmap = pixmap.scaled(self.uploaded_image_width, self.uploaded_image_height, QtCore.Qt.KeepAspectRatio)
            pixmap = pixmap.scaled(self.uploaded_image_width, self.uploaded_image_height)
            # pixmap.scaledToHeight(uploaded_image_height)
            for btn in self.option_buttons:
                btn.setEnabled(True)
                btn.setVisible(True)
            self.uploaded_image.setPixmap(pixmap)
            self.image_upload_button.setVisible(False)
            self.remove_image_button.setVisible(True)
            self.uploaded_image.setVisible(True)
            self.error_label.setText("")
            self.send_message_button_y = self.uploaded_image_y + self.uploaded_image_height + self.h_ten
            self.send_message_button.move(self.send_message_button_x, self.send_message_button_y)
            # self.connected_button_y = self.uploaded_image_y + self.uploaded_image_height + self.h_ten
            # self.connected_button.move(self.connected_button_x, self.connected_button_y)


    def display_side_panel(self, flag):
        self.image_upload_button.setVisible(flag)
        self.send_message_button.setVisible(flag)
        self.text_area.setVisible(flag)
        self.text_area_label.setVisible(flag)
        # for btn in self.option_buttons:
        #     btn.setVisible(flag)


    def disable_side_panel(self):
        self.image_upload_button.setEnabled(False)
        self.remove_image_button.setEnabled(False)
        self.send_message_button.setEnabled(False)
        self.text_area.setEnabled(False)
        # self.connected_button.setEnabled(False)
        if self.image_filename:
            for btn in self.option_buttons:
                btn.setEnabled(False)


    def enable_side_panel(self):
        self.image_upload_button.setEnabled(True)
        self.remove_image_button.setEnabled(True)
        self.send_message_button.setEnabled(True)
        self.text_area.setEnabled(True)
        if self.image_filename:
            for btn in self.option_buttons:
                btn.setEnabled(True)


    def excel_cancel_button_clicked(self):
        self.excel_approve_flag = False
        self.excel_filename = None
        self.excel_data = None
        self.excel_upload_button.setEnabled(True)
        self.excel_approve_button.setEnabled(False)
        self.excel_table.setVisible(False)
        self.excel_approve_button.setVisible(False)
        self.excel_cancel_button.setVisible(False)
        self.disable_side_panel()


    def excel_approve_button_clicked(self):
        self.excel_approve_flag = True
        self.excel_upload_button.setEnabled(False)
        self.excel_approve_button.setEnabled(False)
        self.display_side_panel(True)
        self.enable_side_panel()


    def excel_upload_button_clicked(self):
        self.set_status("", "green")
        self.excel_filename = QtGui.QFileDialog.getOpenFileName(self)
        self.excel_filename = str(self.excel_filename)
        base = os.path.basename(self.excel_filename)
        file, ext = os.path.splitext(base)
        if ext != '.xlsx':
            print ("Invalid Excel File")
            # l.resize(self.upload_button.frameGeometry().width(), self.upload_button.frameGeometry().height())
            # l.move(self.upload_button.frameGeometry().width() + upload_button_x + 10 , upload_button_y)
            self.set_error("Not Valid Excel File", "red")
            self.excel_table.setVisible(False)
            self.excel_approve_button.setVisible(False)
            self.excel_filename = None
        else:
            self.error_label.setText("")
            flag = False
            try:
                ws = gm.openWorkbook(self.excel_filename, sheetname)
                columns, items = gm.getRows(ws)
                flag = True
            except Exception as err:
                # print (str(err))
                self.set_error("Excel file can't be read : currupted ", "red")
                flag = False
            if flag:
                self.excel_data = {"columns": columns, "items": items}
                self.excel_table.clear()
                self.excel_table.setColumnCount(len(columns))
                self.excel_table.setRowCount(len(items))
                # self.excel_table.setHorizontalHeaderLabels(columns)
                for i in range(len(columns)):
                    # print i
                    self.excel_table.setHorizontalHeaderItem(i, QtGui.QTableWidgetItem(columns[i]))
                for i in range(len(items)):
                    item = items[i]
                    for j in range(len(columns)):
                        key = str(self.excel_table.horizontalHeaderItem(j).text())

                        if item[key] is not None:
                            val = QtGui.QTableWidgetItem(item[key])
                        else:
                            val = QtGui.QTableWidgetItem("None")
                        self.excel_table.setItem(i, j, val)

                self.excel_table.setVisible(True)
                self.excel_approve_button.setVisible(True)
                self.excel_approve_button.setEnabled(True)
                self.excel_approve_flag = False
                self.excel_cancel_button.setVisible(True)
                self.excel_cancel_button.setEnabled(True)
                self.tb_label.setVisible(False)
                self.tb.setVisible(False)


    # def connected_button_clicked(self):
    #     # backend_driver.setFlag(self.connected)
    #     self.connected_button.setEnabled(False)
    #     # print "In GUI ", self.connected.value

    def n_finish(self):
        # if self.proc.is_alive():
        #     self.proc.join()
        try:
            if self.go_process.state() == QtCore.QProcess.Running:
                self.go_process.kill()
        except Exception as err:
            print(err)
        self.tb.setVisible(True)
        self.tb_label.setVisible(True)
        self.refresh()
        return

    def refresh(self):
        self.set_error("", "red")
        # self.tb.setVisible(False)
        # self.tb_label.setVisible(False)
        # self.clear_button.setVisible(False)
        self.qr_image.setVisible(False)
        self.excel_upload_button.setEnabled(True)
        self.excel_filename = None
        self.excel_table.setVisible(False)
        self.excel_approve_flag = False
        self.excel_data = None
        self.excel_approve_button.setVisible(False)
        self.excel_cancel_button.setVisible(False)
        self.display_side_panel(False)
        self.refresh_side_panel()


    def refresh_side_panel(self):
        self.text_area.clear()
        self.text_area.setEnabled(True)
        self.image_upload_button.setEnabled(True)
        self.image_filename = None
        self.send_message_button.setEnabled(True)
        # self.connected_button.setVisible(False)
        # self.connected_button.setEnabled(True)
        self.remove_image_button.setVisible(False)
        self.uploaded_image.setVisible(False)
        self.send_message_button_y = self.image_upload_button_y + 30
        self.send_message_button.move(self.send_message_button_x, self.send_message_button_y)
        # self.connected_button_y = self.image_upload_button_y + 30
        # self.connected_button.move(self.connected_button_x, self.connected_button_y)
        for btn in self.option_buttons:
            btn.setVisible(False)


    def send_message_button_clicked(self):
        flag = True
        message = ""
        try:
            message = str(self.text_area.toPlainText())
        except:
            self.set_error("Message contains non-ascii character(s)", "red")
            flag = False

        if flag and len(message) == 0:
            self.set_error("Enter Message", "red")
            flag = False
        # print self.excel_filename
        if flag and self.excel_filename is None:
            self.set_error("Upload Excel Sheet", "red")
            flag = False

        if flag and not self.excel_approve_flag:
            self.set_error("Upload Excel Sheet", "red")
            flag = False

        if flag:
            try:
                # print self.excel_data['items'][0]
                message.format(**self.excel_data['items'][0])
            except Exception as err:
                self.set_error("Column " + str(err) + " Not Found In Excel", "red")
                flag = False
        if flag:
            self.excel_cancel_button.setEnabled(False)
            self.image_upload_button.setEnabled(False)
            self.remove_image_button.setEnabled(False)
            self.send_message_button.setEnabled(False)
            self.text_area.setEnabled(False)

            file = self.excel_filename
            image = self.image_filename

            self.time_stamp = str(datetime.datetime.now()).split('.')[0]
            self.time_stamp = self.time_stamp.replace(" ", "")
            self.time_stamp = self.time_stamp.replace("-", "_")
            self.time_stamp = self.time_stamp.replace(":", "_")

            self.err_dir = os.path.expanduser("~\Desktop")
            self.error_file = os.path.join(self.err_dir, "error.xlsx")

            # error_file = os.path.basename(self.error_file)
            # fnm, ext = os.path.splitext(error_file)
            fnm, ext = "error", ".xlsx"
            xl_filename = fnm + "_" + self.time_stamp + ext
            self.error_file = os.path.join(self.err_dir, xl_filename)

            fnm, ext = "success", ".xlsx"
            success_filename = fnm + "_" + self.time_stamp + ext
            success_filename = os.path.join(self.err_dir, success_filename)

            used_columns = [fn for _, fn, _, _ in Formatter().parse(message) if fn is not None]
            if "name" not in used_columns:
                used_columns.append("name")
            if "phone" not in used_columns:
                used_columns.append("phone")

            not_sent_row = 2
            try:
                wb = xl.load_workbook(self.error_file)
                ws = wb.active
                not_sent_row = len(list(ws.rows)) + 1

                print ('Opening file...of errors', not_sent_row)
            except:
                print ("File not found, creating new file...for errors")
                wb = Workbook()

            go_file = "whatsapp.xlsx"

            if os.path.exists(go_file):
                os.remove(go_file)

            if os.path.exists(self.qr_file):
                os.remove(self.qr_file)

            if os.path.exists("status.txt"):
                os.remove("status.txt")

            go_wb = Workbook()
            go_ws = go_wb.active
            go_ws.title = "temp"

            ws = wb.active
            ws.title = "failed"

            self.write_details(self.error_file, wb, ws, 1, ['phone', 'message', 'reason'])
            self.write_details(go_file, go_wb, go_ws, 1, ['phone', 'message'])
            go_file_count = 2

            not_sent = list()

            items = self.excel_data["items"]
            for item in items:
                try:
                    msg = message.format(**item)
                    key_flag = False
                    for key in used_columns:
                        if item[key] is None:
                            not_sent.append(item['phone'])
                            self.write_details(self.error_file, wb, ws, not_sent_row, [item['phone'], msg, "column " + key + " is null"])
                            not_sent_row = not_sent_row + 1
                            key_flag = True
                            break
                    if key_flag:
                        continue

                except Exception as err:
                    print(err)
                    return
                self.write_details(go_file, go_wb, go_ws, go_file_count, [item['phone'], msg])
                go_file_count += 1

            go_exe = "whatsapp.exe"
            print ("option :", self.selected_option)
            if self.image_filename:
                arguments = [go_file, self.image_filename, str(self.selected_option), self.time_stamp]
            else:
                arguments = [go_file, self.time_stamp]
            self.go_process.start(go_exe, arguments)
            self.notifier.start()
            self.tb.setText('<a target="_blank" href="file:///{path}">{path}</a>'.format(path=self.error_file))


    def go_finished(self):
        pass

    def write_details(self, input_filename, wb, ws, row, message):
        for i in range(len(message)):
            _ = ws.cell(row=row, column=i + 1, value=message[i])
        wb.save(filename=input_filename)


    def set_error(self, msg, color):
        self.error_label.setText(msg)
        self.error_label.setStyleSheet('color: ' + color)
        self.error_label.resize(self.error_label.sizeHint())

    def set_status(self, msg, color):
        self.status_label.setText(msg)
        self.status_label.setStyleSheet('color: ' + color)
        self.status_label.resize(self.status_label.sizeHint())

    def center(self):
        qr = self.frameGeometry()
        cp = QtGui.QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())



def main():
    app = QtGui.QApplication(sys.argv)
    ex = Window()
    sys.exit(app.exec_())

if __name__ == '__main__':
    # freeze_support()
    main()