import openpyxl as xl
import string

def openWorkbook(filename, sheetname):
    wb = xl.load_workbook(filename)
    print wb.sheetnames
    if sheetname in wb.sheetnames:
        ws = wb[sheetname]
        return ws
    else:
        print "Sheet " + sheetname + "does not exist"
        print "Returning active sheet"
        return wb.active

def getRowFromIndex(row):
    # row = list(ws.rows)[index]
    values = list()
    for cell in row:
        values.append(cell.value)
    return values

def getRows(ws):
    # col_n = getAllColumns(ws)
    rows_list = list(ws.rows)
    col_names = getRowFromIndex(rows_list[0])
    values_list = list()
    for i in range(1, len(rows_list)):
        row = getRowFromIndex(rows_list[i])
        row_dict = {}
        for index in range(len(col_names)):
            if row[index] is not None:
                try:
                    row_dict[col_names[index]] = str(row[index])
                except Exception as err:
                    print(err)

            else:
                row_dict[col_names[index]] = None
        values_list.append(row_dict)
    return col_names, values_list
