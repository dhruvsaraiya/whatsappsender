from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoAlertPresentException

from selenium.webdriver.common.keys import Keys

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


import time
import datetime
import os
import re


import openpyxl as xl
from openpyxl import Workbook

import get_messages_v2 as gm

from cStringIO import StringIO
import win32clipboard
from PIL import Image

from string import Formatter

# import sys
# sys.stdout = open('backend_output.txt', 'w')

# import os


# filename = 'whatsapp.xlsx'
sheetname = 'Sheet1'
chrome_driver_path = 'important/chromedriver'
# err_dir = os.path.join(os.environ["HOMEPATH"], "Desktop")
# xl_filename = os.path.join(err_dir, "error.xlsx")
err_dir = os.path.expanduser("~/Desktop")
xl_filename = os.path.join(err_dir, "error.xlsx")
success_filename = os.path.join(err_dir, "success.xlsx")
# image_dir = 'images'

website = "https://web.whatsapp.com/"

## Row number in message file
m=0


def send_to_clipboard(clip_type, data):
    win32clipboard.OpenClipboard()
    win32clipboard.EmptyClipboard()
    win32clipboard.SetClipboardData(clip_type, data)
    win32clipboard.CloseClipboard()


def copyImage(filepath):
    image = Image.open(filepath)
    output = StringIO()
    image.convert("RGB").save(output, "BMP")
    data = output.getvalue()[14:]
    output.close()
    send_to_clipboard(win32clipboard.CF_DIB, data)


def sendImage(driver, image_name):
    copyImage(image_name)
    class_id = '_2S1VP'
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, class_id)))
    msg_element = driver.find_element_by_class_name(class_id)
    msg_element.send_keys(Keys.CONTROL + "v")

    send_button_class = "_3hV1n"
    WebDriverWait(driver, 10).until(EC.presence_of_element_located(
        (By.CLASS_NAME, send_button_class)))
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
        (By.CLASS_NAME, send_button_class)))
    #print "send button available"
    send_button = driver.find_element_by_class_name(send_button_class)
    send_button.click()


def sendMessage(driver, message):
    # class_id ="pluggable-input-compose"
    class_id = "input-container"
    class_id = '_2S1VP'
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, class_id)))
    msg_element = driver.find_element_by_class_name(class_id)

    message = str(message)
    send_to_clipboard(win32clipboard.CF_TEXT, message)
    msg_element.send_keys(Keys.CONTROL + "v")

    send_button_class = "_2lkdt"
    WebDriverWait(driver, 10).until(EC.presence_of_element_located(
        (By.CLASS_NAME, send_button_class)))
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
        (By.CLASS_NAME, send_button_class)))
    send_button = driver.find_element_by_class_name(send_button_class)
    send_button.click()


def sendCaption(driver, message, image_name):
    message = str(message)

    class_id = '_2S1VP'

    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, class_id)))
    msg_element = driver.find_element_by_class_name(class_id)
    send_to_clipboard(win32clipboard.CF_TEXT, message)
    msg_element.send_keys(Keys.CONTROL + "v")

    time.sleep(0.5)

    copyImage(image_name)

    msg_element.send_keys(Keys.CONTROL + "v")

    send_button_class = "_3hV1n"
    WebDriverWait(driver, 10).until(EC.presence_of_element_located(
        (By.CLASS_NAME, send_button_class)))
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
        (By.CLASS_NAME, send_button_class)))
    send_button = driver.find_element_by_class_name(send_button_class)
    send_button.click()
    
    return


# def sendNewLine(driver, message):
#     class_id = "input-container"
#     WebDriverWait(driver, 10).until(EC.presence_of_element_located(
#         (By.CLASS_NAME, class_id)))
#     msg_element = driver.find_element_by_class_name(class_id)
#     msg_element.send_keys(message)
#     #print "waiting for the value - message"
#     ##WebDriverWait(driver, 10).until(EC.text_to_be_present_in_element_value(
#     ##   (By.CLASS_NAME, class_id)))
#     ##print "Wait over"
#     #text_to_be_present_in_element_value
#     #pluggable-input-body copyable-text selectable-text
#     #block-compose
#     send_button_class = "compose-btn-send"
#     WebDriverWait(driver, 10).until(EC.presence_of_element_located(
#         (By.CLASS_NAME, send_button_class)))
#     WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
#         (By.CLASS_NAME, send_button_class)))
#     #print "send button available"
#     send_button = driver.find_element_by_class_name(send_button_class)
#     send_button.click()


def openDriver():
    # driver = webdriver.Chrome(chromium_driver_path)
    # driver = webdriver.Chrome(chrome_driver_path)
    driver = webdriver.Chrome(chrome_driver_path)
    return driver


def visitPage(driver, website):
    driver.get(website)
    print "Driver Title: ", driver.title



def searchReceiver(driver, receiver):

    class_search = 'jN-F5'

    xpath = "//input[contains(@class, '"+class_search+"')]"

    input_element = driver.find_element_by_xpath(xpath)

    WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
        (By.XPATH, xpath)))

    # print ("input ***", input_element)
    try:
        input_element.click()
    except Exception as err:
        # print(err)
        print ("error in clicking")
    # time.sleep(0.5)

    send_to_clipboard(win32clipboard.CF_TEXT, str(receiver))

    # win32clipboard.OpenClipboard()
    # data = win32clipboard.GetClipboardData()
    # win32clipboard.CloseClipboard()
    # print ("number:::::::::::", data)
    time.sleep(.5)

    # input_element.send_keys(Keys.CONTROL + "v")

    # attempt_count = 0
    # while(attempt_count<5):
    #     # input_element.send_keys(str(receiver))

    #     # print("attempting")
    #     input_element.send_keys(Keys.CONTROL + "v")

    #     # time.sleep(1)
        
    #     # print('I am here')
    #     value = str(input_element.get_attribute("value"))
        
    #     # print('I reached here with value : '+value)

    #     if (value==str(receiver)):
    #         # print(receiver + " is correct")
    #         break
    #     else:
    #         attempt_count += 1
    #         print("wrong value found :"+ attempt_count+ "got value :"+ value)
    #         clear_field(driver)
    # # print ("this is the end")




    # # print("VALUE:", value_before)



    input_element.send_keys(Keys.CONTROL + "v")
    input_element.send_keys(Keys.ENTER)

    # value_after = str(input_element.get_attribute("value"))

    # print(value_after)

    # print("VALUE:", str(input_element.get_attribute("value")))

    return


def findInSearchResults(driver, receiver):
    # print receiver
    flag = True
    try:
        xpath = "//div[contains(@class, '_2zCDG')]/span[contains(@class, '_1wjpf')]"
        contact = driver.find_element_by_xpath(xpath)
        text = contact.text
        # print ("TEXT : ", text)
        for r in receiver:
            if not re.search(r, text, re.IGNORECASE):
                flag = False

    except Exception as err:
        flag = False
        # clear_field(driver)
        # print str(err), " in verifying correct element"

    return flag

# def findInSearchResults(driver, receiver):
#     receiver = str(receiver)
#     try:
#         css = 'span[title*="' + receiver +'"]'
#         # print 'waiting'
#         WebDriverWait(driver, 5).until(EC.presence_of_element_located(
#             (By.CSS_SELECTOR, css)))
#         # print 'detected'
#         frnd_element = driver.find_element_by_css_selector(css)
#         # print 'selected'
#     except:
#         receiver = receiver[0:5]+" "+receiver[5:]
#         css = 'span[title*="' + receiver +'"]'
#         WebDriverWait(driver, 1).until(EC.presence_of_element_located(
#             (By.CSS_SELECTOR, css)))
#         frnd_element = driver.find_element_by_css_selector(css)
        
#     flag = False
#     for i in range(5):
#         try:
#             WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
#                 (By.CSS_SELECTOR, css)))
#             #print frnd_element
#             frnd_element.click()
#             break
#         except:
#             print "attempt : ", i+1
#     return flag


def write_details(input_filename, wb, ws, row, message):
    _ = ws.cell(row = row, column = 1, value = message[0])
    _ = ws.cell(row = row, column = 2, value = message[1])
    _ = ws.cell(row = row, column = 3, value = message[2])
    _ = ws.cell(row = row, column = 4, value = message[3])
    _ = ws.cell(row = row, column = 5, value = message[4])
    wb.save(filename = input_filename)


def clear_field(driver):
    try:
        class_search = 'jN-F5'

        xpath = "//input[contains(@class, '" + class_search + "')]"

        input_element = driver.find_element_by_xpath(xpath)

        WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
            (By.XPATH, xpath)))


        # # input_element = driver.find_element_by_xpath("//input[contains(@class, 'jN-F5')]")
        # WebDriverWait(driver, 10).until(EC.presence_of_element_located(
        #     (By.CLASS_NAME, class_search)))

        input_element.click()

        input_element.send_keys(Keys.CONTROL + "a")

        input_element.send_keys(Keys.DELETE)

    except Exception as err:

        print "Could not clear field"
        print "clear field exception ", str(err)
        setClear(1)

        while(True):
            try:
                not_cleared_field = driver.find_element_by_xpath("//input[contains(@class, 'jN-F5')]")
                value_of_search_box = not_cleared_field.get_attribute("value")
                print len(value_of_search_box), str(value_of_search_box), " i was wrong here"
                if len(value_of_search_box) == 0:
                    setClear(0)
                    time.sleep(5)
                    break
                else:
                    setClear(1)
                    time.sleep(5)
            except Exception as e:
                print "ERROR ", e
                setClear(1)
                time.sleep(5)


# def clear_field(driver):
#     class_loader = "_2xarx"
#     class_cross = "_3Burg"
#     flag = False
#     try:
#         # time.sleep(10)
#         WebDriverWait(driver, 10).until(EC.presence_of_element_located(
#             (By.CLASS_NAME, class_cross)))
#         cross_element = driver.find_element_by_class_name(class_cross)
#         WebDriverWait(driver, 5).until(EC.element_to_be_clickable(
#                     (By.CLASS_NAME, class_cross)))
#         cross_element.click()
#         time.sleep(5)
#         # raise
#     except Exception as e:
#         print "cross button not found"
#         try:
#             WebDriverWait(driver, 10).until(EC.presence_of_element_located(
#                 (By.CLASS_NAME, class_loader)))
#             loader_element = driver.find_element_by_class_name(class_loader)
#             WebDriverWait(driver, 5).until(EC.element_to_be_clickable(
#                         (By.CLASS_NAME, class_loader)))
#             loader_element.click()
#             time.sleep(5)
#             # raise
#         except Exception as e:
#             print "Could not clear field"
#             print "clear field exception ", str(e)
#             setClear(1)
#             while(True):
#                 try:
#                     not_cleared_field = driver.find_element_by_xpath("//input[contains(@class, 'jN-F5')]")
#                     value_of_search_box = not_cleared_field.get_attribute("value")
#                     print len(value_of_search_box), str(value_of_search_box), " i was wrong here"
#                     if len(value_of_search_box) == 0:
#                         setClear(0)
#                         time.sleep(5)
#                         break
#                     else:
#                         setClear(1)
#                         time.sleep(5)
#                 except Exception as e:
#                     print "ERROR ", e
#                     setClear(1)
#                     time.sleep(5)
#                 # with open('clear.txt', 'r') as f:
#                 #     c = f.read()
#                 #     if c == '1':
#                 #         time.sleep(5)
#                 # if clear.value == 0:
#                 #     pass
#                 # elif clear.value == 1:
#                 #     time.sleep(5)
#             # raw_input("Clear WhatsApp search box then enter to continue")

# clear = 0 (field is blank OK) , clear = 1 (field is not blank Not OK)
# def setClear(clear,val):
#     with clear.get_lock():
#         clear.value = val
#     print "ok", clear.value

def setClear(val):
    with open('clear.txt', 'w') as f:
        f.write(str(val))

def setFlag(connected):
    with connected.get_lock():
        connected.value = 1
    print "ok", connected.value

# def setStatus(status, value):
#     with status.get_lock():
#         status.value = value
#     print "status changed : ", status.value

def setStatus(val):
    with open('status.txt', 'w') as f:
        f.write(str(val))

def setMessageError(val):
    with open('err_msg.txt', 'w') as f:
        f.write(str(val))

# def before_exit(self, driver):
#     driver.quit()
#     sys.exit(0)

def sendMessages(columns, items, input_message, image_path, selected_option, connected, time_stamp):
    with connected.get_lock():
        connected.value = 0

    driver = openDriver()

    driver.implicitly_wait(5)
    visitPage(driver, website)
    setStatus('0,0')
    setClear('0')
    setMessageError('')

    used_columns = [fn for _, fn, _, _ in Formatter().parse(input_message) if fn is not None]

    global xl_filename
    global success_filename

    err_dir = os.path.expanduser("~/Desktop")
    # xl_filename = os.path.join(err_dir, "error.xlsx")

    error_file = os.path.basename(xl_filename)
    fnm, ext = os.path.splitext(error_file)
    xl_filename = fnm + "_" + time_stamp + ext
    xl_filename = xl_filename.replace(" ", "")
    xl_filename = xl_filename.replace("-", "_")
    xl_filename = xl_filename.replace(":", "_")
    xl_filename = os.path.join(err_dir, xl_filename)

    s_file = os.path.basename(success_filename)
    fnm, ext = os.path.splitext(s_file)
    success_filename = fnm + "_" + time_stamp + ext
    success_filename = success_filename.replace(" ", "")
    success_filename = success_filename.replace("-", "_")
    success_filename = success_filename.replace(":", "_")
    success_filename = os.path.join(err_dir, success_filename)
    # try:
    #     os.remove(success_filename)
    # except OSError:
    #     pass

    while(True):
        if connected.value == 1:
            not_sent_row = 2
            sent_row = 2
            try:
                wb = xl.load_workbook(xl_filename)
                ws = wb.active
                not_sent_row = len(list(ws.rows)) + 1

                print 'Opening file...of errors', not_sent_row
            except:
                print "File not found, creating new file...for errors"
                wb = Workbook()

            try:
                s_wb = xl.load_workbook(success_filename)
                s_ws = s_wb.active
                sent_row = len(list(s_ws.rows)) + 1

                print 'Opening file...of success', sent_row
            except:
                print "File not found, creating new file...of success"
                s_wb = Workbook()

            s_ws = s_wb.active
            s_ws.title = "sent"

            ws = wb.active
            ws.title = "failed"

            write_details(xl_filename, wb, ws, 1, ['name', 'phone', 'message', 'reason', 'time'])
            write_details(success_filename, s_wb, s_ws, 1, ['name', 'phone', 'message', 'reason', 'time'])

            sent_phone_no = list()
            not_sent = list()

            for item in items:
                try:
                    message = input_message.format(**item)
                    key_flag = False
                    for key in used_columns:
                        # print "key : ", key, item[key]
                        if item[key] is None:
                            not_sent.append(item['phone'])
                            write_details(xl_filename, wb, ws, not_sent_row, [item['name'], item['phone'], message, "column " + key + " is null", str(datetime.datetime.now())])
                            not_sent_row = not_sent_row + 1
                            key_flag = True
                            break
                    if key_flag:
                        continue

                except Exception as err:
                    err_msg = str(err)
                    setMessageError(err_msg)
                    driver.quit()
                    return

                try:
                    if item['phone'] in sent_phone_no:
                        print "Repeated user: ", item['name'], " with number: ", item['phone']

                    # search_query = str(item['name']) + " " + str(item['phone'])
                    search_query = str(item['phone'])
                    searchReceiver(driver, str(search_query))
                except:
                    not_sent.append(item['phone'])
                    print "Receiver not found: ", item['name']
                    write_details(xl_filename, wb, ws, not_sent_row, [item['name'], item['phone'], message, "Receiver not found: " + item['name'], str(datetime.datetime.now())])
                    not_sent_row = not_sent_row + 1
                    clear_field(driver)
                    # time.sleep(2)
                    continue

                try:
                    # verify_query = [str(item['name']), str(item['phone'])]
                    phone = str(item['phone']).strip()
                    verify_query = [phone[:5], phone[5:]]
                    s_flag = findInSearchResults(driver, verify_query)

                    if not s_flag:
                        print "Wrong element selected for " + item['name']
                        not_sent.append(item['phone'])
                        write_details(xl_filename, wb, ws, not_sent_row, [item['name'], item['phone'], message, "Wrong element selected for " + item['name'], str(datetime.datetime.now())])
                        not_sent_row = not_sent_row + 1
                        clear_field(driver)
                        # time.sleep(2)
                        continue

                    ####################################################
                    
                    if image_path:
                        
                        if selected_option == 0:
                            sendMessage(driver, message)
                            time.sleep(1)
                            sendImage(driver, image_path)
                        
                        elif selected_option == 1:
                            sendImage(driver, image_path)
                            time.sleep(1)
                            sendMessage(driver, message)
                        
                        elif selected_option == 2:
                            sendCaption(driver, message, image_path)
                        
                        else:
                            # print ("no option")
                            pass
                    else:
                        sendMessage(driver, message)
                    ####################################################
                    # print "Number: ", item['phone'], " message sent: ", message
                    sent_phone_no.append(item['phone'])

                    write_details(success_filename, s_wb, s_ws, sent_row, [item['name'], item['phone'], message, "Message sent", str(datetime.datetime.now())])
                    sent_row += 1

                    # print (float(len(sent_phone_no)) / float(len(items)))
                    status_value = str(int((float(len(sent_phone_no)) / float(len(items))) * 100.0))
                    status_value += ","
                    status_value += str(int((float(len(not_sent)) / float(len(items))) * 100.0))
                    # print "status completed : ", status_value, len(sent_phone_no)
                    setStatus(status_value)
                except TimeoutException:
                    print "Timeout for receiver ", item['name']
                    not_sent.append(item['phone'])
                    write_details(xl_filename, wb, ws, not_sent_row, [item['name'], item['phone'], message, "Timeout for receiver " + item['name'], str(datetime.datetime.now())])
                    not_sent_row += 1
                    # clear_field(driver)
                # time.sleep(1)

            while(True):
                try:
                    not_sent_message = driver.find_element_by_xpath("//div[@class='_1VfKB']/span[@data-icon='status-time']")
                    # print "messeages not sent till now "
                except NoSuchElementException as err:
                    print "messeages sent"
                    status_value = str(int((float(len(sent_phone_no)) / float(len(items))) * 100.0))
                    status_value += ","
                    status_value += str(int((float(len(not_sent)) / float(len(items))) * 100.0))
                    setStatus(status_value)
                    break
            break
        else:
            time.sleep(3)
    print("process finished part 2 started")

    # read success file

    ws = gm.openWorkbook(success_filename, "sent")
    s_columns, s_items = gm.getRows(ws)

    phone_index = list()
    for i in items:
        phone_index.append(str(i['phone']))
    for s_item in s_items:
        try:
            index_of_dict = phone_index.index(str(s_item['phone']))
            del items[index_of_dict]
            del phone_index[index_of_dict]
        except Exception as err:
            pass
            # print (err)

    try:
        sendMessages_2(driver, columns, items, input_message, image_path, selected_option, connected, time_stamp)
        driver.quit()
    except Exception as err:
        print(err)

    return


###################################################################

def sendImage_2(driver, image_name):
    copyImage(image_name)
    class_id = '_2S1VP'
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, class_id)))
    msg_element = driver.find_element_by_class_name(class_id)
    msg_element.send_keys(Keys.CONTROL + "a")
    msg_element.send_keys(Keys.DELETE)
    msg_element.send_keys(Keys.CONTROL + "v")

    send_button_class = "_3hV1n"
    WebDriverWait(driver, 10).until(EC.presence_of_element_located(
        (By.CLASS_NAME, send_button_class)))
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
        (By.CLASS_NAME, send_button_class)))
    #print "send button available"
    send_button = driver.find_element_by_class_name(send_button_class)
    send_button.click()


def sendMessage_2(driver, message):
    # class_id ="pluggable-input-compose"
    class_id = "input-container"
    class_id = '_2S1VP'
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, class_id)))
    msg_element = driver.find_element_by_class_name(class_id)

    message = str(message)
    send_to_clipboard(win32clipboard.CF_TEXT, message)
    msg_element.send_keys(Keys.CONTROL + "a")
    msg_element.send_keys(Keys.CONTROL + "v")

    send_button_class = "_2lkdt"
    WebDriverWait(driver, 10).until(EC.presence_of_element_located(
        (By.CLASS_NAME, send_button_class)))
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
        (By.CLASS_NAME, send_button_class)))
    #print "send button available"
    send_button = driver.find_element_by_class_name(send_button_class)
    send_button.click()


def sendCaption_2(driver, message, image_name):
    message = str(message)

    class_id = '_2S1VP'

    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, class_id)))
    msg_element = driver.find_element_by_class_name(class_id)
    send_to_clipboard(win32clipboard.CF_TEXT, message)
    msg_element.send_keys(Keys.CONTROL + "a")
    msg_element.send_keys(Keys.CONTROL + "v")

    time.sleep(0.5)

    copyImage(image_name)

    msg_element.send_keys(Keys.CONTROL + "v")

    send_button_class = "_3hV1n"
    WebDriverWait(driver, 10).until(EC.presence_of_element_located(
        (By.CLASS_NAME, send_button_class)))
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
        (By.CLASS_NAME, send_button_class)))
    send_button = driver.find_element_by_class_name(send_button_class)
    send_button.click()
    return


def visitPage_2(driver, website):
    driver.get(website)
    try:
        driver.switch_to_alert().accept()
    except NoAlertPresentException:
        pass

def setFlag_2(connected, value):
    with connected.get_lock():
        connected.value = value
    print "Connected :", connected.value


def do_temp(driver, phone, tmp_msg):

    api_site = "https://api.whatsapp.com/send?phone=91{}&text={}".format(phone, tmp_msg)
    visitPage_2(driver, api_site)
    try:

        xpath = "//a[contains(@class, 'button button--simple button--primary')]"
        send_button = driver.find_element_by_xpath(xpath)
        send_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, xpath)))
        send_button.click()

    except Exception as e:
        print "ERROR ", e
        driver.quit()

    # time.sleep(5)
    return


def sendMessages_2(driver, columns, items, input_message, image_path, selected_option, connected, time_stamp):
    # visitPage(driver, website)

    setStatus('0,0')
    setMessageError('')

    used_columns = [fn for _, fn, _, _ in Formatter().parse(input_message) if fn is not None]

    global xl_filename
    global success_filename

    # err_dir = os.path.expanduser("~/Desktop")
    # # xl_filename = os.path.join(err_dir, "error.xlsx")

    # error_file = os.path.basename(xl_filename)
    # fnm, ext = os.path.splitext(error_file)
    # xl_filename = fnm + "_" + time_stamp + ext
    # xl_filename = xl_filename.replace(" ", "")
    # xl_filename = xl_filename.replace("-", "_")
    # xl_filename = xl_filename.replace(":", "_")
    # xl_filename = os.path.join(err_dir, xl_filename)

    # s_file = os.path.basename(success_filename)
    # fnm, ext = os.path.splitext(s_file)
    # success_filename = fnm + "_" + time_stamp + ext
    # success_filename = success_filename.replace(" ", "")
    # success_filename = success_filename.replace("-", "_")
    # success_filename = success_filename.replace(":", "_")
    # success_filename = os.path.join(err_dir, success_filename)

    not_sent_row = 2
    sent_row = 2
    try:
        wb = xl.load_workbook(xl_filename)
        ws = wb.active
        not_sent_row = len(list(ws.rows)) + 1

        print 'Opening file...of errors', not_sent_row
    except:
        print "File not found, creating new file...for errors"
        wb = Workbook()

    try:
        s_wb = xl.load_workbook(success_filename)
        s_ws = s_wb.active
        sent_row = len(list(s_ws.rows)) + 1

        print 'Opening file...of success', sent_row
    except:
        print "File not found, creating new file...of success"
        s_wb = Workbook()

    s_ws = s_wb.active
    s_ws.title = "sent"

    ws = wb.active
    ws.title = "failed"

    sent_phone_no = list()
    not_sent = list()

    for item in items:
        try:
            message = input_message.format(**item)
            key_flag = False
            for key in used_columns:
                # print "key : ", key, item[key]
                if item[key] is None:
                    not_sent.append(item['phone'])
                    write_details(xl_filename, wb, ws, not_sent_row, [item['name'], item['phone'], message, "In 2nd Iteration...... column " + key + " is null", str(datetime.datetime.now())])
                    not_sent_row = not_sent_row + 1
                    key_flag = True
                    break
            if key_flag:
                continue

        except Exception as err:
            err_msg = str(err)
            setMessageError(err_msg)
            driver.quit()
            return

        tmp_msg = "greetings from logiqids!"
        do_temp(driver, item['phone'].strip(), tmp_msg)

        while(True):
            if connected.value == 1:
                try:
                    # wait for page to be loaded
                    try:
                        xpath = "//img[contains(@class, 'Qgzj8')]"
                        contact = WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH, xpath)))
                    except TimeoutException as err:
                        print("....still continue if found further....")
                    except Exception as err:
                        print("Profile found")

                    ####################################################

                    if image_path:
                        
                        if selected_option == 0:
                            sendMessage_2(driver, message)
                            time.sleep(1)
                            sendImage_2(driver, image_path)
                        
                        elif selected_option == 1:
                            sendImage_2(driver, image_path)
                            time.sleep(1)
                            sendMessage_2(driver, message)
                        
                        elif selected_option == 2:
                            sendCaption_2(driver, message, image_path)
                        
                        else:
                            # print ("no option")
                            pass
                    else:
                        sendMessage_2(driver, message)
                    ####################################################
                    # print "Number: ", item['phone'], " message sent: ", message
                    sent_phone_no.append(item['phone'])

                    write_details(success_filename, s_wb, s_ws, sent_row, [item['name'], item['phone'], message, "In 2nd Iteration..... Message sent", str(datetime.datetime.now())])
                    sent_row += 1

                    # print (float(len(sent_phone_no)) / float(len(items)))
                    status_value = str(int((float(len(sent_phone_no)) / float(len(items))) * 100.0))
                    status_value += ","
                    status_value += str(int((float(len(not_sent)) / float(len(items))) * 100.0))
                    # print "status completed : ", status_value, len(sent_phone_no)
                    setStatus(status_value)
                except TimeoutException:
                    print "Timeout for receiver ", item['name']
                    not_sent.append(item['phone'])
                    write_details(xl_filename, wb, ws, not_sent_row, [item['name'], item['phone'], message, "In 2nd iteration.... Timeout for receiver " + item['name'], str(datetime.datetime.now())])
                    not_sent_row += 1
                    # clear_field(driver)
                # time.sleep(1)
                break
            else:
                time.sleep(3)

    while(True):
        try:
            not_sent_message = driver.find_element_by_xpath("//div[@class='_1VfKB']/span[@data-icon='status-time']")
            # print "messeages not sent till now "
        except NoSuchElementException as err:
            print "messeage sent"
            break
    print("process finished with part 2")
    driver.quit()
    return